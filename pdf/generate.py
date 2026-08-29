"""
Phil - PDF generation (certificates and session records).

Uses reportlab directly rather than a template engine, matching the approach
already used in the project's build_certificate_pdf.py / build_session_record_pdf.py
scripts, so a future production build can port this straight into the small
internal PDF microservice the Technical Build Spec recommends (section 2.1),
reusing this logic rather than rewriting it.
"""

import os
import datetime


# Below this many completed follow-ups the impact report gives counts but no
# percentage. A proportion from two or three courses reads as a finding when it
# is really a coincidence, and a head presenting it would be caught out.
FOLLOWUP_MIN_FOR_SHARE = 5


# The same wording the mentor picked from, so a report never shows a bare
# number whose meaning the reader has to guess.
MOOD_LABELS = {1: "very low", 2: "low", 3: "mixed", 4: "settled", 5: "bright"}
ENGAGEMENT_LABELS = {1: "wouldn't take part", 2: "took part reluctantly",
                     3: "took part when prompted", 4: "took part willingly",
                     5: "led it themselves"}


def _rating_word(value, labels):
    try:
        return labels.get(int(value), "not rated")
    except (TypeError, ValueError):
        return "not rated"


def uk(value):
    """2026-09-07 as 07/09/2026. Anything unparseable comes back untouched, so a
    half-filled record still prints rather than raising mid-document."""
    if not value:
        return ""
    try:
        return datetime.date.fromisoformat(str(value)[:10]).strftime("%d/%m/%Y")
    except (TypeError, ValueError):
        return str(value)


def uk_long(value):
    """7 September 2026. For the certificate, which a child keeps: slashes read
    as paperwork where a written month reads as an occasion."""
    if not value:
        return ""
    try:
        d = datetime.date.fromisoformat(str(value)[:10])
    except (TypeError, ValueError):
        return str(value)
    return "%d %s %d" % (d.day, d.strftime("%B"), d.year)


def today_uk():
    return datetime.date.today().strftime("%d/%m/%Y")
import re
import math

import body_map
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.lib.colors import HexColor
from reportlab.pdfgen import canvas
from reportlab.lib.utils import simpleSplit

NAVY = HexColor("#1B2A4A")
TEAL = HexColor("#1D9E75")
TEAL_DARK = HexColor("#0F6E56")
# The book cover in the Phil mark. The masthead uses it so the band and the logo
# are the same green rather than two greens.
TEAL_DARKER = HexColor("#085041")
AMBER = HexColor("#EF9F27")
CREAM = HexColor("#FBF8F2")
INK = HexColor("#2C2C2A")
MUTED = HexColor("#5F5E5A")
RED = HexColor("#A32D2D")
CARD = HexColor("#FFFFFF")
BORDER = HexColor("#E4E1D6")

# Generated PDFs must live on the same persistent volume as the database. The
# previous location, <repo>/data/pdfs, is inside the container image, so every
# certificate and session record was destroyed on the next deploy while the
# database kept a pdf_path pointing at the missing file — which is what produced
# "Certificate not found". PHIL_PDF_DIR overrides; otherwise sit beside the DB.
def _pdf_dir():
    explicit = os.environ.get("PHIL_PDF_DIR")
    if explicit:
        return explicit
    db_path = os.environ.get("PHIL_DB_PATH")
    if db_path:
        return os.path.join(os.path.dirname(db_path), "pdfs")
    return os.path.join(os.path.dirname(os.path.dirname(__file__)), "data", "pdfs")


PDF_DIR = _pdf_dir()
os.makedirs(PDF_DIR, exist_ok=True)


def _phil_mark(c, x, y, size=13 * mm):
    """Draws the Phil mark: a green book with a cream P, a page behind it and a
    coral bookmark. Redrawn in reportlab primitives rather than embedded as an
    image, so it stays crisp at any size and needs no asset file on the volume.

    Coordinates are proportional to size, taken from the SVG the app uses, so
    the two can't drift apart."""
    u = size / 120.0  # the source artwork is on a 120-unit grid

    # Page behind the book
    c.setFillColor(HexColor("#EAE2CC"))
    c.roundRect(x + 30 * u, y + 14 * u, 66 * u, 90 * u, 4 * u, fill=1, stroke=0)

    # Bookmark
    c.setFillColor(HexColor("#D85A30"))
    c.rect(x + 62 * u, y + 104 * u, 7 * u, 14 * u, fill=1, stroke=0)

    # Book cover
    c.setFillColor(HexColor("#0F6E56"))
    c.roundRect(x + 30 * u, y + 20 * u, 56 * u, 84 * u, 6 * u, fill=1, stroke=0)

    # The P as a glyph rather than a transcribed path. Converting the SVG's
    # bezier curves by hand meant flipping the y-axis, and getting that subtly
    # wrong produced a shape that wasn't the letter at all. A font draws it
    # correctly at any size.
    c.setFillColor(HexColor("#F3EFE4"))
    glyph = 62 * u
    c.setFont("Helvetica-Bold", glyph)
    # Optically centred on the cover: the cap sits slightly above the baseline
    # centre, so nudge down rather than using the exact midpoint.
    c.drawCentredString(x + 58 * u, y + 46 * u, "P")


def _doc_header(c, w, h, margin, doc_title, meta=None, subtitle="Structured mentoring for schools"):
    """The masthead every Phil report shares.

    A report that gets tabled at a governors' meeting or sent to a parent has to
    look like it came from somewhere. That means a consistent masthead, the
    document's name stated plainly, and the identifying details set out as
    labelled pairs rather than run together in a sentence."""
    y = h - margin

    # Masthead in the logo's own deep green, so the band and the mark belong to
    # each other rather than sitting side by side in different palettes.
    # At 13mm the mark was legible on screen but muddy in print, so the band is
    # taller and the mark larger.
    band_h = 25 * mm
    band_y = y - 6 * mm
    c.setFillColor(TEAL_DARKER)
    c.rect(0, band_y, w, band_h, fill=1, stroke=0)
    # A cream hairline along the bottom edge, picking up the page colour in the
    # mark, so the band reads as designed rather than as a block of colour.
    c.setStrokeColor(HexColor("#EAE2CC"))
    c.setLineWidth(1.2)
    c.line(0, band_y, w, band_y)

    _phil_mark(c, margin, band_y + 5 * mm, size=15 * mm)

    text_x = margin + 20 * mm
    c.setFillColor(HexColor("#FFFFFF"))
    c.setFont("Helvetica-Bold", 15)
    c.drawString(text_x, band_y + 12.5 * mm, "Phil")
    c.setFont("Helvetica", 7.5)
    c.setFillColor(HexColor("#A7D9C8"))
    c.drawString(text_x, band_y + 8 * mm, subtitle.upper())

    c.setFillColor(HexColor("#EAE2CC"))
    c.setFont("Helvetica-Bold", 9)
    c.drawRightString(w - margin, band_y + 10.5 * mm, doc_title.upper())

    y -= 23 * mm

    # Identifying details as labelled pairs, two per row.
    if meta:
        c.setFont("Helvetica", 9)
        col_w = (w - 2 * margin) / 2
        for i, (label, value) in enumerate(meta):
            col = i % 2
            if col == 0 and i:
                y -= 5.5 * mm
            cx = margin + col * col_w
            c.setFillColor(MUTED)
            c.drawString(cx, y, f"{label}:")
            c.setFillColor(INK)
            c.setFont("Helvetica-Bold", 9)
            c.drawString(cx + c.stringWidth(f"{label}:", "Helvetica", 9) + 2 * mm, y, str(value))
            c.setFont("Helvetica", 9)
        y -= 7 * mm

    c.setStrokeColor(TEAL_DARKER)
    c.setLineWidth(1.2)
    c.line(margin, y, w - margin, y)
    return y - 9 * mm


def _doc_footer(c, w, margin, page_no, note="Confidential. Share only with those who need it."):
    """Page number and a confidentiality line, on every page of every report."""
    c.setStrokeColor(BORDER)
    c.setLineWidth(0.5)
    c.line(margin, 14 * mm, w - margin, 14 * mm)
    c.setFillColor(MUTED)
    c.setFont("Helvetica", 7.5)
    c.drawString(margin, 10 * mm, note)
    c.drawRightString(w - margin, 10 * mm, f"Page {page_no}")


def _doc_section(c, x, y, label, max_width):
    """A section heading with a hairline under it, so sections actually read as
    sections rather than as bold text in a wall."""
    c.setFillColor(TEAL_DARKER)
    c.setFont("Helvetica-Bold", 10.5)
    c.drawString(x, y, label.upper())
    y -= 2.5 * mm
    c.setStrokeColor(BORDER)
    c.setLineWidth(0.5)
    c.line(x, y, x + max_width, y)
    return y - 5 * mm


GOLD = HexColor("#B8912F")
GOLD_LIGHT = HexColor("#E4C86B")
GOLD_PALE = HexColor("#F5EAC6")


def _award_rosette(c, cx, cy, r):
    """A gold rosette with two ribbon tails, drawn in reportlab primitives.

    Certificates carry one because people expect one; a child who has finished
    five weeks of this should get something that looks like an award rather than
    a receipt."""
    # Ribbon tails first, so the medallion sits over them.
    c.setFillColor(GOLD)
    for direction in (-1, 1):
        p = c.beginPath()
        p.moveTo(cx + direction * r * 0.42, cy - r * 0.25)
        p.lineTo(cx + direction * r * 0.80, cy - r * 1.75)
        p.lineTo(cx + direction * r * 0.38, cy - r * 1.45)
        p.lineTo(cx + direction * r * 0.05, cy - r * 1.85)
        p.lineTo(cx + direction * r * 0.02, cy - r * 0.25)
        p.close()
        c.setFillColor(GOLD if direction < 0 else GOLD_LIGHT)
        c.drawPath(p, fill=1, stroke=0)

    # Fluted edge: a ring of small circles reads as a rosette at any size.
    c.setFillColor(GOLD_LIGHT)
    points = 12
    for i in range(points):
        a = 2 * math.pi * i / points
        c.circle(cx + math.cos(a) * r * 0.82, cy + math.sin(a) * r * 0.82,
                 r * 0.30, fill=1, stroke=0)

    c.setFillColor(GOLD)
    c.circle(cx, cy, r * 0.86, fill=1, stroke=0)
    c.setFillColor(GOLD_PALE)
    c.circle(cx, cy, r * 0.66, fill=1, stroke=0)
    c.setStrokeColor(GOLD)
    c.setLineWidth(0.8)
    c.circle(cx, cy, r * 0.52, fill=0, stroke=1)

    # A star at the centre, rather than text that would fight the wordmark.
    c.setFillColor(GOLD)
    star = c.beginPath()
    for i in range(10):
        a = math.pi / 2 + i * math.pi / 5
        rad = r * 0.42 if i % 2 == 0 else r * 0.17
        x, y = cx + math.cos(a) * rad, cy + math.sin(a) * rad
        star.moveTo(x, y) if i == 0 else star.lineTo(x, y)
    star.close()
    c.drawPath(star, fill=1, stroke=0)


def certificate_pdf(pupil_name, course_title, issued_date, enrolment_id,
                    establishment_name=None, mentor_name=None, module_number=None):
    """A certificate a pupil would be happy to take home and a school happy to file.

    Modelled on the conventions of an awarding-body certificate: portrait, serif,
    the issuing body named at the top, a unique reference, and signature lines.
    Those details are what make a certificate read as a record rather than a
    printout — the reference in particular, because it implies something was
    written down somewhere."""
    path = os.path.join(PDF_DIR, f"certificate_{enrolment_id}.pdf")
    w, h = A4  # portrait, as awarding bodies use
    c = canvas.Canvas(path, pagesize=(w, h))

    # Cream ground rather than white: warmer in the hand, and the colour the
    # rest of Phil uses.
    c.setFillColor(HexColor("#FDFBF4"))
    c.rect(0, 0, w, h, fill=1, stroke=0)

    # A green band at the head, so the page is recognisably Phil's before a word
    # is read. Then a double rule: heavy green outer, gold hairline inner —
    # restraint reads as official, ornament reads as a template.
    margin = 16 * mm
    c.setFillColor(TEAL_DARKER)
    c.rect(0, h - 12 * mm, w, 12 * mm, fill=1, stroke=0)
    c.setFillColor(GOLD)
    c.rect(0, h - 13.2 * mm, w, 1.2 * mm, fill=1, stroke=0)

    c.setStrokeColor(TEAL_DARKER)
    c.setLineWidth(2.2)
    c.rect(margin, margin, w - 2 * margin, h - 2 * margin - 10 * mm, fill=0, stroke=1)
    c.setStrokeColor(GOLD_LIGHT)
    c.setLineWidth(0.8)
    c.rect(margin + 3 * mm, margin + 3 * mm, w - 2 * margin - 6 * mm,
           h - 2 * margin - 10 * mm - 6 * mm, fill=0, stroke=1)

    y = h - 42 * mm

    # The Phil mark, then the issuing body — the same book used everywhere else.
    _phil_mark(c, w / 2 - 7 * mm, y - 2 * mm, size=14 * mm)
    y -= 10 * mm

    c.setFillColor(TEAL_DARKER)
    c.setFont("Times-Bold", 22)
    c.drawCentredString(w / 2, y, "PHIL")
    y -= 6 * mm
    c.setFillColor(MUTED)
    c.setFont("Times-Roman", 9.5)
    c.drawCentredString(w / 2, y, "STRUCTURED MENTORING FOR SCHOOLS")

    y -= 6 * mm
    c.setStrokeColor(GOLD)
    c.setLineWidth(0.8)
    c.line(w / 2 - 32 * mm, y, w / 2 + 32 * mm, y)

    y -= 18 * mm
    c.setFillColor(NAVY)
    c.setFont("Times-Bold", 26)
    c.drawCentredString(w / 2, y, "Certificate of Completion")

    y -= 20 * mm
    c.setFillColor(INK)
    c.setFont("Times-Roman", 12)
    c.drawCentredString(w / 2, y, "This is to certify that")

    y -= 16 * mm
    c.setFillColor(NAVY)
    c.setFont("Times-Bold", 28)
    c.drawCentredString(w / 2, y, _clean_pdf_text(pupil_name))

    # A rule under the name, as a certificate would have.
    y -= 4 * mm
    name_width = max(c.stringWidth(_clean_pdf_text(pupil_name), "Times-Bold", 28), 60 * mm)
    c.setStrokeColor(BORDER)
    c.setLineWidth(0.6)
    c.line(w / 2 - name_width / 2 - 8 * mm, y, w / 2 + name_width / 2 + 8 * mm, y)

    if establishment_name:
        y -= 10 * mm
        c.setFillColor(MUTED)
        c.setFont("Times-Roman", 11.5)
        c.drawCentredString(w / 2, y, f"of {_clean_pdf_text(establishment_name)}")

    y -= 16 * mm
    c.setFillColor(INK)
    c.setFont("Times-Roman", 12)
    c.drawCentredString(w / 2, y, "has successfully completed the five-session mentoring course")

    y -= 16 * mm
    c.setFillColor(TEAL_DARK)
    c.setFont("Times-Bold", 19)
    c.drawCentredString(w / 2, y, _clean_pdf_text(course_title))

    if module_number:
        y -= 8 * mm
        c.setFillColor(MUTED)
        c.setFont("Times-Roman", 10)
        c.drawCentredString(w / 2, y, f"Module {str(module_number).zfill(2)} of the Phil course library")

    # The rosette sits between the course title and the signatures, in the empty
    # middle of the page — the place the eye lands.
    _award_rosette(c, w / 2, margin + 92 * mm, 12 * mm)

    # Signature lines. An unsigned line still says a person stands behind this.
    sig_y = margin + 62 * mm
    left_x, right_x = w / 2 - 62 * mm, w / 2 + 12 * mm
    c.setStrokeColor(BORDER)
    c.setLineWidth(0.7)
    c.setFillColor(MUTED)
    c.setFont("Times-Roman", 9)
    if establishment_name:
        # Two signatories: the mentor who ran it, and the school that stands
        # behind it.
        c.line(left_x, sig_y, left_x + 50 * mm, sig_y)
        c.line(right_x, sig_y, right_x + 50 * mm, sig_y)
        c.drawString(left_x, sig_y - 5 * mm, "Mentor")
        c.drawString(right_x, sig_y - 5 * mm, "Signed on behalf of the school")
        name_x = left_x + 1 * mm
    else:
        # An individual mentor has no school to sign for it, so one centred line
        # rather than a second line nobody can sign.
        centre_x = w / 2 - 25 * mm
        c.line(centre_x, sig_y, centre_x + 50 * mm, sig_y)
        c.drawCentredString(w / 2, sig_y - 5 * mm, "Mentor")
        name_x = centre_x + 1 * mm
    if mentor_name:
        c.setFillColor(INK)
        c.setFont("Times-Italic", 11)
        c.drawString(name_x, sig_y + 2.5 * mm, _clean_pdf_text(mentor_name))

    # Date and reference, footer left and right, as on a real certificate.
    foot_y = margin + 38 * mm
    c.setFillColor(MUTED)
    c.setFont("Times-Roman", 10)
    c.drawString(left_x, foot_y, f"Date of issue: {uk_long(issued_date)}")
    c.drawRightString(right_x + 50 * mm, foot_y,
                      f"Certificate no. PHL-{str(enrolment_id).zfill(6)}")

    c.setFont("Times-Roman", 8)
    c.setFillColor(MUTED)
    c.drawCentredString(w / 2, margin + 14 * mm,
                        "This certificate records completion of a mentoring course. "
                        "It is not a formal qualification.")

    c.showPage()
    c.save()
    return path


def _wrap(c, text, x, y, max_width, font="Helvetica", size=9, leading=12, color=INK):
    c.setFont(font, size)
    c.setFillColor(color)
    lines = simpleSplit(text or "", font, size, max_width)
    for line in lines:
        c.drawString(x, y, line)
        y -= leading
    return y


def session_record_pdf(record, enrolment, pupil_name, course_title, week_title, mentor_name,
                       resource_work=None):
    """
    record: sqlite3.Row from session_records
    """
    path = os.path.join(PDF_DIR, f"session_{record['id']}.pdf")
    w, h = A4
    c = canvas.Canvas(path, pagesize=A4)
    margin = 18 * mm
    x = margin
    y = h - margin

    max_width = w - 2 * margin
    y = _doc_header(c, w, h, margin, "Session record", meta=[
        ("Pupil", pupil_name),
        ("Course", course_title),
        ("Session", week_title),
        ("Date", uk(record["date"])),
        ("Mentor", mentor_name),
    ])

    def section(label, text):
        nonlocal y
        y = _doc_section(c, x, y, label, max_width)
        c.setFillColor(INK)
        y = _wrap(c, text or "-", x, y, max_width)
        y -= 5 * mm

    c.setFillColor(INK)
    c.setFont("Helvetica", 10)
    mood = _rating_word(record["mood_rating"], MOOD_LABELS)
    engagement = _rating_word(record["engagement_rating"], ENGAGEMENT_LABELS)
    c.drawString(x, y, f"Mood: {mood}    Took part: {engagement}")
    y -= 8 * mm

    section("What happened", record["what_happened"])
    section("Reflect", record["reflection_goal"])
    section("Summary for this session", record["mentor_notes"])
    # What the pupil actually wrote, under the resource's own name. Resources
    # used but not written on are listed separately, so the record distinguishes
    # "we used this" from "here is what came of it".
    written_on = {title for title, lines in (resource_work or []) if lines}
    for title, lines in (resource_work or []):
        if lines:
            section(title, "\n".join(lines))
    others = [r.strip() for r in (record["resources_used"] or "").split(",")
              if r.strip() and r.strip() not in written_on]
    if others:
        section("Other resources used", ", ".join(others))

    # Safeguarding block, always rendered, matching the mandatory-step convention
    # established across the project's other PDF exports.
    flagged = bool(record["safeguarding_flag"])
    label = "Safeguarding: flagged" if flagged else "Safeguarding: not flagged this session"
    disclaimer = ("This record is not a safeguarding report and Phil takes no action on it. "
                  "The mentor remains responsible for following their establishment's own "
                  "safeguarding procedure.")
    note_text = record["safeguarding_note"] or "-"

    # The box used to be a fixed 26mm whatever it contained, so a longer note
    # pushed the disclaimer out through the bottom edge. Measure first, then draw
    # a box that fits.
    inner_w = max_width - 8 * mm
    note_lines = simpleSplit(note_text, "Helvetica", 9, inner_w)
    disc_lines = simpleSplit(disclaimer, "Helvetica-Oblique", 7.5, inner_w)
    box_h = (9 * mm                      # label
             + len(note_lines) * 4.6 * mm
             + 3 * mm                     # gap
             + len(disc_lines) * 3.2 * mm
             + 5 * mm)                    # bottom padding
    box_h = max(box_h, 22 * mm)

    if y - box_h < margin:
        c.showPage()
        y = h - margin

    c.setFillColor(HexColor("#FCEBEB") if flagged else HexColor("#E1F5EE"))
    c.rect(x, y - box_h, max_width, box_h, fill=1, stroke=0)
    c.setStrokeColor(RED if flagged else TEAL)
    c.setLineWidth(1.2)
    c.rect(x, y - box_h, max_width, box_h, fill=0, stroke=1)

    c.setFillColor(RED if flagged else TEAL_DARK)
    c.setFont("Helvetica-Bold", 11)
    c.drawString(x + 4 * mm, y - 7 * mm, label)

    c.setFillColor(INK)
    c.setFont("Helvetica", 9)
    text_y = y - 12 * mm
    for line in note_lines:
        c.drawString(x + 4 * mm, text_y, line)
        text_y -= 4.6 * mm

    c.setFillColor(MUTED)
    c.setFont("Helvetica-Oblique", 7.5)
    text_y -= 3 * mm
    for line in disc_lines:
        c.drawString(x + 4 * mm, text_y, line)
        text_y -= 3.2 * mm

    _doc_footer(c, w, margin, 1)
    c.showPage()
    c.save()
    return path


def session_summaries_pdf(enrolment_id, pupil_name, course_title, mentor_name, rows):
    """The five session summaries on one sheet, for writing the course summary.

    Session 6 asks the mentor to review the course before writing anything. Left
    to themselves that means opening five records in five tabs, so most wouldn't.
    One sheet they can put beside them is the difference between the instruction
    being followed and ignored."""
    path = os.path.join(PDF_DIR, f"summaries_{enrolment_id}.pdf")
    w, h = A4
    c = canvas.Canvas(path, pagesize=A4)
    margin = 18 * mm
    x = margin
    max_width = w - 2 * margin

    y = _doc_header(c, w, h, margin, "Session summaries", meta=[
        ("Pupil", pupil_name),
        ("Course", course_title),
        ("Mentor", mentor_name),
        ("Printed", today_uk()),
    ])

    c.setFillColor(MUTED)
    c.setFont("Helvetica-Oblique", 9)
    y = _wrap(c, "For writing the course summary and next steps in session 6. Read these in order before you start.",
              x, y, max_width, size=9)
    y -= 6 * mm

    page_no = 1
    if not rows:
        c.setFillColor(MUTED)
        c.setFont("Helvetica-Oblique", 10)
        c.drawString(x, y, "No sessions recorded yet.")

    for row in rows:
        # Each block now carries what happened, the goal, the pupil's own work
        # and any safeguarding note, so the old 34mm estimate would start a
        # session near the foot of a page and split it.
        block = 34 * mm
        block += 6 * mm * len([p for p in (row["what_happened"] or "").split("\n\n") if p.strip()]) \
            if "what_happened" in row.keys() else 0
        block += 10 * mm * len(row["resource_work"] or []) \
            if "resource_work" in row.keys() else 0
        block += 8 * mm if ("safeguarding_note" in row.keys()
                            and (row["safeguarding_note"] or "").strip()) else 0
        block = min(block, 150 * mm)
        if y - block < margin + 24 * mm:
            _doc_footer(c, w, margin, page_no)
            c.showPage()
            page_no += 1
            y = _doc_header(c, w, h, margin, "Session summaries",
                            meta=[("Pupil", pupil_name), ("Course", course_title)])

        y = _doc_section(c, x, y, f"Session {row['week_number']} \u00b7 {row['week_title']}", max_width)
        c.setFillColor(MUTED)
        c.setFont("Helvetica", 8.5)
        mood = _rating_word(row["mood_rating"], MOOD_LABELS)
        engagement = _rating_word(row["engagement_rating"], ENGAGEMENT_LABELS)
        c.drawString(x, y, f"{uk(row['date'])}    Mood: {mood}    Took part: {engagement}"
                           + ("    Safeguarding flagged" if row["safeguarding_flag"] else ""))
        y -= 6 * mm

        # The summary is the point of the sheet, so it leads.
        c.setFillColor(INK)
        summary = (row["mentor_notes"] or "").strip()
        if summary:
            y = _wrap(c, summary, x, y, max_width, size=10)
        else:
            c.setFillColor(MUTED)
            c.setFont("Helvetica-Oblique", 9.5)
            c.drawString(x, y, "No summary written for this session.")
            y -= 5 * mm

        # What actually happened in the room. Stored as "Check-in: ... / Input:
        # ... / Activity: ...", and until now it lived only in the session
        # record, so this sheet could not replace opening one.
        happened = (row["what_happened"] or "").strip() if "what_happened" in row.keys() else ""
        for part in [p.strip() for p in happened.split("\n\n") if p.strip()]:
            y -= 1.5 * mm
            c.setFillColor(INK)
            y = _wrap(c, part, x, y, max_width, size=9.5)

        goal = (row["reflection_goal"] or "").strip()
        if goal:
            y -= 2 * mm
            c.setFillColor(TEAL_DARK)
            c.setFont("Helvetica-Bold", 9)
            c.drawString(x, y, "Goal set:")
            c.setFillColor(INK)
            y = _wrap(c, goal, x + 18 * mm, y, max_width - 18 * mm, size=9.5)

        # What the pupil ticked or wrote on their own sheets.
        work = row["resource_work"] if "resource_work" in row.keys() else None
        for name, lines in (work or []):
            y -= 2 * mm
            c.setFillColor(TEAL_DARK)
            c.setFont("Helvetica-Bold", 9)
            c.drawString(x, y, name)
            y -= 4.5 * mm
            c.setFillColor(INK)
            for line in lines:
                y = _wrap(c, line, x + 5 * mm, y, max_width - 5 * mm, size=9)

        # The flag alone told a reader something was raised but not what.
        note = (row["safeguarding_note"] or "").strip() if "safeguarding_note" in row.keys() else ""
        if note:
            y -= 2 * mm
            c.setFillColor(RED)
            c.setFont("Helvetica-Bold", 9)
            c.drawString(x, y, "Safeguarding:")
            c.setFillColor(INK)
            y = _wrap(c, note, x + 24 * mm, y, max_width - 24 * mm, size=9.5)
        y -= 7 * mm

    _doc_footer(c, w, margin, page_no)
    c.showPage()
    c.save()
    return path


def impact_report_pdf(establishment_id, establishment_name, f):
    """What mentoring achieved across a school, on one page.

    Written to be read by someone who wasn't involved: a governor, an SLT link,
    an inspector. Every figure says what it is based on, and where the evidence
    is thin the report says so rather than filling the space."""
    path = os.path.join(PDF_DIR, f"impact_report_{establishment_id}.pdf")
    w, h = A4
    c = canvas.Canvas(path, pagesize=A4)
    margin = 18 * mm
    x = margin
    max_width = w - 2 * margin

    # The period belongs in the header, not a footnote. A report that does not
    # say what it covers gets quoted as though it covers everything.
    if f.get("date_from") and f.get("date_to"):
        period = f"{uk(f['date_from'])} to {uk(f['date_to'])}"
    elif f.get("date_from"):
        period = f"from {uk(f['date_from'])}"
    elif f.get("date_to"):
        period = f"up to {uk(f['date_to'])}"
    else:
        period = "All activity to date"
    y = _doc_header(c, w, h, margin, "Impact report", meta=[
        ("School", establishment_name),
        ("Period", period),
        ("Issued", today_uk()),
    ])

    def stat(label, value, note=""):
        """One figure, its label, and what it rests on."""
        nonlocal y
        c.setFillColor(TEAL_DARKER)
        c.setFont("Helvetica-Bold", 17)
        c.drawString(x, y, str(value))
        vw = c.stringWidth(str(value), "Helvetica-Bold", 17)
        c.setFillColor(INK)
        c.setFont("Helvetica", 10)
        c.drawString(x + vw + 4 * mm, y + 1, label)
        if note:
            c.setFillColor(MUTED)
            c.setFont("Helvetica", 8.5)
            c.drawString(x + vw + 4 * mm, y - 4.5 * mm, note)
            y -= 4.5 * mm
        y -= 9 * mm

    y = _doc_section(c, x, y, "Reach", max_width)
    stat("pupils supported", f.get("pupils") or 0,
         f"{f.get('enrolments') or 0} course enrolments in total")
    stat("sessions delivered", f.get("sessions") or 0)
    completed = f.get("completed") or 0
    enrolments = f.get("enrolments") or 0
    rate = f"{round(completed / enrolments * 100)}%" if enrolments else "\u2014"
    stat("courses completed", completed,
         f"{rate} of enrolments, with {f.get('active') or 0} still running")
    y -= 2 * mm

    y = _doc_section(c, x, y, "Change", max_width)
    for key, label in (("engagement", "engagement"), ("mood", "mood")):
        n = f.get(f"{key}_n") or 0
        change = f.get(f"{key}_change")
        if n == 0:
            c.setFillColor(MUTED)
            c.setFont("Helvetica-Oblique", 9.5)
            c.drawString(x, y, f"Not enough {label} ratings yet to show change.")
            y -= 7 * mm
            continue
        arrow = "+" if change and change > 0 else ""
        stat(f"average change in {label}", f"{arrow}{change:.1f}",
             f"across {n} course{'' if n == 1 else 's'} rated at the start and again later; "
             f"{f.get(f'{key}_improved') or 0} improved")
    c.setFillColor(MUTED)
    c.setFont("Helvetica-Oblique", 8.5)
    y = _wrap(c, "Ratings are the mentor's judgement at the time, on a 1\u20135 scale. "
                 "Only courses rated more than once are counted. Figures cover "
                 "sessions delivered in the period above, so a course spanning "
                 "two terms appears in both.", x, y, max_width, size=8.5)
    y -= 6 * mm

    y = _doc_section(c, x, y, "Follow-through", max_width)
    stat("course summaries written", f.get("plans_written") or 0,
         "the session 6 write-up other staff can pick up and use")

    # What the follow-up chats found. Below FOLLOWUP_MIN_FOR_SHARE the counts are
    # shown without a proportion: "100%" off two courses is a figure that falls
    # apart the moment a governor asks how many that was, and the honest answer
    # is that it is too early to say.
    done = f.get("followups_done") or 0
    due = f.get("followups_due") or 0
    if due:
        stat("follow-up chats completed", f"{done} of {due}",
             "a sit-down with the pupil a few weeks after the course ended")
    if done:
        sustained = f.get("followups_sustained") or 0
        helped = f.get("followups_helped") or 0
        if done >= FOLLOWUP_MIN_FOR_SHARE:
            stat("behaviour no longer showing", f"{sustained} of {done}",
                 f"{round(sustained * 100 / done)}% of courses followed up")
            stat("mentor judged the course helped", f"{helped} of {done}",
                 f"{round(helped * 100 / done)}% rated better or some change")
        else:
            stat("behaviour no longer showing", f"{sustained} of {done}",
                 "too few follow-ups yet to give a meaningful percentage")
            stat("mentor judged the course helped", f"{helped} of {done}",
                 "rated better or some change")

    overdue = f.get("reviews_overdue") or 0
    stat("follow-up chats overdue", overdue,
         "past the date agreed with the pupil" if overdue else "nothing outstanding")
    stat("sessions with a safeguarding note", f.get("safeguarding") or 0,
         "recorded by mentors; Phil takes no action on these")
    y -= 2 * mm

    courses = f.get("courses") or []
    if courses:
        if y < margin + 45 * mm:
            _doc_footer(c, w, margin, 1)
            c.showPage()
            y = _doc_header(c, w, h, margin, "Impact report",
                            meta=[("School", establishment_name)])
        y = _doc_section(c, x, y, "Courses used", max_width)
        c.setFont("Helvetica", 9.5)
        shown = 0
        for row in courses:
            if shown >= 14 or y < margin + 25 * mm:
                break
            count_text = f"{row['n']} enrolled \u00b7 {row['completed']} completed"
            # Measured, not counted: leave room for the figures on the right so a
            # long course title can't run into them.
            room = max_width - c.stringWidth(count_text, "Helvetica", 9.5) - 8 * mm
            title = row["title"]
            if c.stringWidth(title, "Helvetica", 9.5) > room:
                while title and c.stringWidth(title + "\u2026", "Helvetica", 9.5) > room:
                    title = title[:-1]
                title += "\u2026"
            c.setFillColor(INK)
            c.drawString(x, y, title)
            c.setFillColor(MUTED)
            c.drawRightString(w - margin, y, count_text)
            y -= 5.5 * mm
            shown += 1
        # Say so rather than stopping silently: a school using every course
        # would otherwise lose six of them with no sign they existed.
        if shown < len(courses):
            remaining = len(courses) - shown
            c.setFillColor(MUTED)
            c.setFont("Helvetica-Oblique", 9)
            c.drawString(x, y, f"and {remaining} more course"
                               f"{'' if remaining == 1 else 's'} \u2014 the full list is "
                               "in the mentoring list report")
            y -= 5.5 * mm

    _doc_footer(c, w, margin, 1)
    c.showPage()
    c.save()
    return path


def pupil_report_pdf(pupil_id, pupil_name, establishment_name, courses, period="All time"):
    """Everything a pupil has done, across every course.

    The per-course report answers "how did that course go". This answers "what
    do I need to know about this child" — the question a new form tutor, a
    SENCO, or the next school actually asks. Each course contributes its support
    plan, so the plans sit together rather than one per file."""
    path = os.path.join(PDF_DIR, f"pupil_report_{pupil_id}.pdf")
    w, h = A4
    c = canvas.Canvas(path, pagesize=A4)
    margin = 18 * mm
    x = margin
    max_width = w - 2 * margin

    meta = [("Pupil", pupil_name)]
    if establishment_name:
        meta.append(("School", establishment_name))
    meta += [("Covering", period), ("Courses", len(courses)), ("Issued", today_uk())]
    y = _doc_header(c, w, h, margin, "Pupil report", meta=meta)
    page_no = 1

    if not courses:
        c.setFillColor(MUTED)
        c.setFont("Helvetica-Oblique", 10)
        c.drawString(x, y, "No courses recorded for this pupil yet."
                     if period == "All time" else
                     f"No courses ran for this pupil in {period.lower()}.")
        _doc_footer(c, w, margin, page_no)
        c.showPage()
        c.save()
        return path

    for course in courses:
        if y < margin + 60 * mm:
            _doc_footer(c, w, margin, page_no)
            c.showPage()
            page_no += 1
            y = _doc_header(c, w, h, margin, "Pupil report",
                            meta=[("Pupil", pupil_name), ("Continued", "")])

        y = _doc_section(c, x, y, course["title"], max_width)
        c.setFillColor(MUTED)
        c.setFont("Helvetica", 8.5)
        bits = [f"Mentor: {course['mentor_name']}", f"Started {uk(course['start_date'])}"]
        bits.append("Completed" if course["status"] == "completed"
                    else f"In progress, session {course['current_week']}")
        if course.get("sessions_recorded") is not None:
            bits.append(f"{course['sessions_recorded']} sessions recorded")
        c.drawString(x, y, "    ".join(bits))
        y -= 6 * mm

        # Session by session. Without this the report answered "which courses"
        # but not "what happened in them", which is the question a new form
        # tutor or a SENCO is actually asking.
        def _page_break():
            nonlocal page_no
            _doc_footer(c, w, margin, page_no)
            c.showPage()
            page_no += 1
            yy = _doc_header(c, w, h, margin, "Pupil report",
                             meta=[("Pupil", pupil_name), ("Continued", "")])
            return _doc_section(c, x, yy, course["title"] + " (continued)", max_width)

        for s in (course.get("sessions") or []):
            if s.get("staff_only"):
                continue                      # the write-up prints as the summary below
            if y < margin + 34 * mm:
                y = _page_break()
            c.setFillColor(INK)
            c.setFont("Helvetica-Bold", 9.5)
            c.drawString(x, y, f"Session {s['week_number']}: {s['title']}   {uk(s.get('date', ''))}")
            y -= 4.6 * mm
            if (s.get("what_happened") or "").strip():
                y = _wrap(c, s["what_happened"].strip(), x + 4 * mm, y, max_width - 4 * mm,
                          size=9, leading=12, color=INK)
            if (s.get("reflection_goal") or "").strip():
                y -= 0.5 * mm
                y = _wrap(c, "Goal set: " + s["reflection_goal"].strip(), x + 4 * mm, y,
                          max_width - 4 * mm, size=9, leading=12, color=MUTED)
            for title, lines in (s.get("resource_work") or []):
                if y < margin + 22 * mm:
                    y = _page_break()
                c.setFillColor(MUTED)
                c.setFont("Helvetica-Bold", 8)
                c.drawString(x + 4 * mm, y, title.upper())
                y -= 4 * mm
                for line in lines:
                    y = _wrap(c, line, x + 8 * mm, y, max_width - 8 * mm, size=8.5,
                              leading=11, color=INK)
            # A flag is the thing a reader must not miss, so it is stated even
            # when the note itself is empty.
            if s.get("safeguarding_flag"):
                if y < margin + 20 * mm:
                    y = _page_break()
                note = (s.get("safeguarding_note") or "").strip()
                c.setFillColor(RED)
                c.setFont("Helvetica-Bold", 8.5)
                c.drawString(x + 4 * mm, y, "SAFEGUARDING CONCERN RECORDED")
                y -= 4.2 * mm
                if note:
                    y = _wrap(c, note, x + 8 * mm, y, max_width - 8 * mm, size=9,
                              leading=12, color=INK)
            y -= 4 * mm

        plan = (course.get("support_plan") or "").strip()
        if plan:
            c.setFillColor(TEAL_DARK)
            c.setFont("Helvetica-Bold", 9)
            c.drawString(x, y, "Course summary and next steps")
            y -= 5 * mm
            c.setFillColor(INK)
            y = _wrap(c, plan, x, y, max_width, size=9.5)
        else:
            c.setFillColor(MUTED)
            c.setFont("Helvetica-Oblique", 9.5)
            c.drawString(x, y, "No course summary written for this course yet.")
            y -= 5 * mm

        # The follow-up is the part that shows whether the course held. A report
        # that stops at the certificate says what was delivered, not what
        # changed, and the second is the question being asked.
        fu = course.get("follow_up")
        if fu:
            y -= 1 * mm
            c.setFillColor(TEAL_DARK)
            c.setFont("Helvetica-Bold", 9)
            c.drawString(x, y, f"Follow-up chat, {uk(fu.get('date', ''))}")
            y -= 5 * mm
            c.setFillColor(INK)
            c.setFont("Helvetica", 9)
            c.drawString(x, y, "    ".join([
                f"Helped: {fu.get('helped_label', '')}",
                f"Behaviour: {fu.get('behaviour_label', '')}",
                f"Next: {fu.get('next_step_label', '')}",
            ]))
            y -= 5 * mm
            voice = (fu.get("pupil_voice") or "").strip()
            if voice:
                c.setFillColor(INK)
                y = _wrap(c, f"In their words: \u201c{voice}\u201d", x, y, max_width, size=9.5)
            note = (fu.get("next_step_note") or "").strip()
            if note:
                c.setFillColor(MUTED)
                y = _wrap(c, note, x, y, max_width, size=9)
        elif course["status"] == "completed":
            c.setFillColor(MUTED)
            c.setFont("Helvetica-Oblique", 9.5)
            c.drawString(x, y, "No follow-up chat recorded yet.")
            y -= 5 * mm

        flagged = course.get("safeguarding_count") or 0
        if fu and fu.get("safeguarding_flag"):
            flagged += 1
        if flagged:
            y -= 1 * mm
            c.setFillColor(RED)
            c.setFont("Helvetica-Bold", 8.5)
            c.drawString(x, y, f"{flagged} record{'' if flagged == 1 else 's'} carried a "
                               "safeguarding note. See the session records.")
            y -= 5 * mm
        y -= 6 * mm

    _doc_footer(c, w, margin, page_no)
    c.showPage()
    c.save()
    return path


def mentee_report_pdf(enrolment_id, pupil_name, course_title, mentor_name, start_date,
                       current_week, status, weeks, reflection=None, support_plan=None):
    """
    weeks: list of dicts with keys week_number, title, objective, date (session date recorded)
                (only pass this when the viewer is entitled to see it, per spec 7.5a/7.6)
    """
    path = os.path.join(PDF_DIR, f"mentee_report_{enrolment_id}.pdf")
    w, h = A4
    c = canvas.Canvas(path, pagesize=A4)
    margin = 18 * mm
    x = margin
    y = h - margin
    max_width = w - 2 * margin

    y = _doc_header(c, w, h, margin, "Course report", meta=[
        ("Pupil", pupil_name),
        ("Course", course_title),
        ("Mentor", mentor_name),
        ("Started", uk(start_date)),
        ("Status", "Completed" if status == "completed" else f"Week {current_week} of 5"),
        ("Issued", today_uk()),
    ])

    page_no = 1
    y = _doc_section(c, x, y, "Sessions covered", max_width)

    for wk in weeks:
        if y < margin + 30 * mm:
            _doc_footer(c, w, margin, page_no)
            c.showPage()
            page_no += 1
            y = _doc_header(c, w, h, margin, "Course report",
                            meta=[("Pupil", pupil_name), ("Course", course_title)])
            y = _doc_section(c, x, y, "Sessions covered (continued)", max_width)
        c.setFillColor(INK)
        c.setFont("Helvetica-Bold", 10.5)
        c.drawString(x, y, f"Week {wk['week_number']}: {wk['title']}  ({uk(wk.get('date',''))})")
        y -= 5 * mm
        y = _wrap(c, wk.get("objective", ""), x + 4 * mm, y, max_width - 4 * mm, size=9.5)
        y -= 3 * mm

        # What the pupil ticked and wrote on the sheets that session. A body map
        # filled in during session 1 was only ever visible in that session's own
        # download, so it never reached the report a SENCO or a new form tutor
        # actually reads.
        for title, lines in (wk.get("resource_work") or []):
            if y < margin + 24 * mm:
                _doc_footer(c, w, margin, page_no)
                c.showPage()
                page_no += 1
                y = _doc_header(c, w, h, margin, "Course report",
                                meta=[("Pupil", pupil_name), ("Course", course_title)])
                y = _doc_section(c, x, y, "Sessions covered (continued)", max_width)
            c.setFillColor(MUTED)
            c.setFont("Helvetica-Bold", 8.5)
            c.drawString(x + 4 * mm, y, title.upper())
            y -= 4.2 * mm
            for line in lines:
                y = _wrap(c, line, x + 8 * mm, y, max_width - 8 * mm, size=9,
                          leading=12, color=INK)
            y -= 2 * mm
        y -= 3 * mm

    if not weeks:
        c.setFillColor(MUTED)
        c.setFont("Helvetica-Oblique", 10)
        c.drawString(x, y, "No sessions recorded yet.")
        y -= 6 * mm

    if support_plan:
        if y < margin + 45 * mm:
            _doc_footer(c, w, margin, page_no)
            c.showPage()
            page_no += 1
            y = _doc_header(c, w, h, margin, "Course report",
                            meta=[("Pupil", pupil_name), ("Course", course_title)])
        y -= 3 * mm
        y = _doc_section(c, x, y, "Course summary and next steps", max_width)
        c.setFillColor(INK)
        y = _wrap(c, support_plan, x, y, max_width, size=9.5)

    _doc_footer(c, w, margin, page_no)
    c.showPage()
    c.save()
    return path


def full_mentoring_report_pdf(title, entries, out_name):
    """
    entries: list of dicts, each with pupil_name, course_title, mentor_name, start_date,
             current_week, status, weeks (list)
    One section per entry, all in a single PDF, for a whole-establishment bulk export
    or a single named pupil. Same restricted fields as the individual course report
    (Enrolment, Course, Week, SessionRecord.date only), plus CompletionReflection
    where the enrolment is completed, per spec section 7.7a.
    """
    path = os.path.join(PDF_DIR, f"{out_name}.pdf")
    w, h = A4
    c = canvas.Canvas(path, pagesize=A4)
    margin = 18 * mm
    max_width = w - 2 * margin
    x = margin
    y = h - margin

    y = _doc_header(c, w, h, margin, "Mentoring report", meta=[
        ("Report", title),
        ("Issued", today_uk()),
        ("Enrolments", len(entries)),
    ])

    if not entries:
        c.setFillColor(MUTED)
        c.setFont("Helvetica-Oblique", 11)
        c.drawString(x, y, "No enrolments to report.")
        _doc_footer(c, w, margin, 1)
        c.showPage()
        c.save()
        return path

    for entry in entries:
        if y < margin + 40 * mm:
            c.showPage()
            y = h - margin

        c.setStrokeColor(BORDER)
        c.line(x, y, w - margin, y)
        y -= 8 * mm

        c.setFillColor(TEAL_DARK)
        c.setFont("Helvetica-Bold", 13)
        c.drawString(x, y, f"{entry['pupil_name']}  -  {entry['course_title']}")
        y -= 6 * mm

        c.setFillColor(MUTED)
        c.setFont("Helvetica", 9.5)
        status_label = "Completed" if entry["status"] == "completed" else f"Week {entry['current_week']} of 5"
        c.drawString(x, y, f"Mentor: {entry['mentor_name']}    Started: {uk(entry['start_date'])}    Status: {status_label}")
        y -= 8 * mm

        for wk in entry["weeks"]:
            if y < margin + 20 * mm:
                c.showPage()
                y = h - margin
            c.setFillColor(INK)
            c.setFont("Helvetica-Bold", 10)
            c.drawString(x, y, f"Week {wk['week_number']}: {wk['title']}  ({uk(wk.get('date',''))})")
            y -= 5 * mm
            y = _wrap(c, wk.get("objective", ""), x + 4 * mm, y, max_width - 4 * mm, size=9)
            y -= 3 * mm

        if not entry["weeks"]:
            c.setFillColor(MUTED)
            c.setFont("Helvetica-Oblique", 9.5)
            c.drawString(x, y, "No sessions recorded yet.")
            y -= 6 * mm

        y -= 6 * mm

    c.showPage()
    c.save()
    return path


def caseload_report_xlsx(rows, show_mentor_col, out_name, period="All time"):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    path = os.path.join(PDF_DIR, f"{out_name}.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Mentoring list"

    headers = ["Pupil", "Course"]
    if show_mentor_col:
        headers.append("Mentor")
    headers += ["Started", "Expected end", "Progress", "Certificate", "Follow-up"]

    ws.append([f"Mentoring list \u2014 covering {period}"])
    ws.append([])
    ws.append(headers)
    header_font = Font(bold=True, name="Arial")
    header_fill = PatternFill(start_color="F2EFE6", end_color="F2EFE6", fill_type="solid")
    ws["A1"].font = Font(bold=True, name="Arial", size=12)
    for cell in ws[3]:
        cell.font = header_font
        cell.fill = header_fill

    for r in rows:
        row = [r["pupil"], r["course"]]
        if show_mentor_col:
            row.append(r.get("mentor", ""))
        row += [r["started"], r["scheduled_end"], r["progress"], r["certificate"],
                r.get("follow_up", "-")]
        ws.append(row)

    for col_cells in ws.columns:
        length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 12), 40)

    for row in ws.iter_rows(min_row=4):
        for cell in row:
            cell.font = Font(name="Arial")

    wb.save(path)
    return path


def caseload_report_pdf(title, rows, show_mentor_col, out_name, period="All time"):
    """
    rows: list of dicts with pupil, course, mentor (optional), started, scheduled_end,
          progress, certificate, follow_up
    """
    path = os.path.join(PDF_DIR, f"{out_name}.pdf")
    w, h = landscape(A4)
    c = canvas.Canvas(path, pagesize=(w, h))
    margin = 14 * mm
    x = margin
    y = h - margin

    # The period goes in the header of every report. A PDF found in a drawer in
    # three years is unreadable without it, and "All time" is worth saying too.
    y = _doc_header(c, w, h, margin, "Mentoring list", meta=[
        ("Report", title),
        ("Covering", period),
        ("Issued", today_uk()),
        ("Pupils", len(rows)),
    ])

    # Columns are weighted, not equal: a course title needs three times the room
    # of a date. Equal widths plus a fixed 28-character cut meant long titles
    # ("Safeguarding: exploitation and county lines awareness") ran straight
    # through the next column.
    cols = [("Pupil", 2.0), ("Course", 3.4)]
    if show_mentor_col:
        cols.append(("Mentor", 2.0))
    cols += [("Started", 1.3), ("Expected end", 1.5), ("Progress", 1.6),
             ("Certificate", 1.3), ("Follow-up", 1.9)]
    total_weight = sum(weight for _, weight in cols)
    avail = w - 2 * margin
    widths = [avail * weight / total_weight for _, weight in cols]

    def fit(text, width, font, size):
        """Trim to what actually fits, measured, with an ellipsis if trimmed."""
        text = str(text)
        if c.stringWidth(text, font, size) <= width:
            return text
        while text and c.stringWidth(text + "\u2026", font, size) > width:
            text = text[:-1]
        return text + "\u2026"

    def draw_row(values, bold=False, fill=None):
        nonlocal y
        if fill:
            c.setFillColor(fill)
            c.rect(x, y - 6 * mm, avail, 7 * mm, fill=1, stroke=0)
        font = "Helvetica-Bold" if bold else "Helvetica"
        c.setFont(font, 8.5)
        c.setFillColor(NAVY if bold else INK)
        cx = x
        for val, cw in zip(values, widths):
            c.drawString(cx + 2, y - 4.5 * mm, fit(val, cw - 6, font, 8.5))
            cx += cw
        y -= 7 * mm

    draw_row([label for label, _ in cols], bold=True, fill=HexColor("#F2EFE6"))
    c.setStrokeColor(BORDER)
    c.line(x, y + 2, w - margin, y + 2)

    for r in rows:
        if y < margin + 10 * mm:
            c.showPage()
            y = h - margin
            draw_row([label for label, _ in cols], bold=True, fill=HexColor("#F2EFE6"))
        values = [r["pupil"], r["course"]]
        if show_mentor_col:
            values.append(r.get("mentor", ""))
        values += [r["started"], r["scheduled_end"], r["progress"], r["certificate"],
                   r.get("follow_up", "-")]
        draw_row(values)

    c.showPage()
    c.save()
    return path


def _clean_pdf_text(text):
    if not text:
        return ""
    return text.replace("\uf0b7", "-").replace("\ue0b7", "-")


_STEP_RE = re.compile(r"^(\d+)[.\)]?\s+(.+)$")


def _parse_numbered_steps(body):
    """Pull out numbered lines like '1 Trigger - something happens' or
    '2. Craving builds: the brain signals it' into short node labels plus
    the full line for a legend underneath the diagram."""
    steps = []
    for line in (body or "").split("\n"):
        m = _STEP_RE.match(line.strip())
        if not m:
            continue
        num, rest = m.group(1), m.group(2)
        short = re.split(r"\s[-:]\s|:\s", rest, maxsplit=1)[0].strip()
        if len(short) > 22:
            short = short[:20].rstrip() + "..."
        steps.append({"num": num, "short": short, "full": f"{num}. {rest}"})
    return steps


def _draw_cycle_diagram(c, cx, cy, radius, steps, color=TEAL_DARK):
    """Draws a circular cycle diagram: numbered nodes arranged in a ring,
    connected by arrows showing the cycle repeating."""
    n = len(steps)
    if n == 0:
        return
    node_r = 8.5 * mm
    for i, step in enumerate(steps):
        angle = math.pi / 2 - (2 * math.pi * i / n)
        step["_pos"] = (cx + radius * math.cos(angle), cy + radius * math.sin(angle))

    c.setStrokeColor(color)
    c.setLineWidth(1.4)
    for i in range(n):
        x1, y1 = steps[i]["_pos"]
        x2, y2 = steps[(i + 1) % n]["_pos"]
        dx, dy = x2 - x1, y2 - y1
        dist = math.hypot(dx, dy)
        if dist == 0:
            continue
        ux, uy = dx / dist, dy / dist
        sx, sy = x1 + ux * node_r * 1.35, y1 + uy * node_r * 1.35
        ex, ey = x2 - ux * node_r * 1.35, y2 - uy * node_r * 1.35
        c.line(sx, sy, ex, ey)
        ah = 2.6 * mm
        line_angle = math.atan2(ey - sy, ex - sx)
        left = (ex - ah * math.cos(line_angle - math.radians(28)), ey - ah * math.sin(line_angle - math.radians(28)))
        right = (ex - ah * math.cos(line_angle + math.radians(28)), ey - ah * math.sin(line_angle + math.radians(28)))
        p = c.beginPath()
        p.moveTo(ex, ey)
        p.lineTo(*left)
        p.lineTo(*right)
        p.close()
        c.setFillColor(color)
        c.drawPath(p, fill=1, stroke=0)

    for step in steps:
        nx, ny = step["_pos"]
        c.setFillColor(color)
        c.circle(nx, ny, node_r, fill=1, stroke=0)
        c.setFillColor(CARD)
        c.setFont("Helvetica-Bold", 10)
        c.drawCentredString(nx, ny - 3.3, step["num"])
        c.setFillColor(INK)
        c.setFont("Helvetica-Bold", 7.3)
        label_y = ny - node_r - 4.2 * mm if ny <= cy + 1 else ny + node_r + 2.4 * mm
        c.drawCentredString(nx, label_y, step["short"])


_SCALE_RE = re.compile(r"^(\d+)\s+(.+)$")


def _parse_scale_points(body):
    points = []
    for line in (body or "").split("\n"):
        m = _SCALE_RE.match(line.strip())
        if m:
            points.append({"num": int(m.group(1)), "label": m.group(2).strip()})
    return points


def _draw_scale_diagram(c, x, y, width, points, color_low=TEAL, color_high=RED):
    """Draws a horizontal numbered scale/thermometer bar, low to high,
    colour-graded, with each point's label beneath its tick."""
    if not points:
        return
    points = sorted(points, key=lambda p: p["num"])
    n = len(points)
    bar_h = 9 * mm
    seg_w = width / n
    for i, pt in enumerate(points):
        t = i / max(1, n - 1)
        col = HexColor(_blend_hex(color_low, color_high, t))
        c.setFillColor(col)
        c.rect(x + i * seg_w, y - bar_h, seg_w, bar_h, fill=1, stroke=0)
    c.setStrokeColor(BORDER)
    c.setLineWidth(0.75)
    c.rect(x, y - bar_h, width, bar_h, fill=0, stroke=1)
    for i, pt in enumerate(points):
        cx_ = x + i * seg_w + seg_w / 2
        c.setFillColor(CARD if i / max(1, n - 1) > 0.35 else INK)
        c.setFont("Helvetica-Bold", 8.5)
        c.drawCentredString(cx_, y - bar_h / 2 - 3, str(pt["num"]))
    label_y = y - bar_h - 5 * mm
    c.setFillColor(INK)
    c.setFont("Helvetica", 6.6)
    for i, pt in enumerate(points):
        cx_ = x + i * seg_w + seg_w / 2
        label = pt["label"]
        if len(label) > 14:
            label = label[:12].rstrip() + "..."
        c.drawCentredString(cx_, label_y, label)


def _blend_hex(c1, c2, t):
    r1, g1, b1 = c1.red, c1.green, c1.blue
    r2, g2, b2 = c2.red, c2.green, c2.blue
    r = r1 + (r2 - r1) * t
    g = g1 + (g2 - g1) * t
    b = b1 + (b2 - b1) * t
    return "#%02X%02X%02X" % (int(r * 255), int(g * 255), int(b * 255))


HEADER_FILL = HexColor("#F2EFE6")


def _cell_lines(text, font, size, w):
    return simpleSplit(str(text), font, size, max(w, 4))


def _draw_grid_table(c, x, top_y, max_width, headers, rows, row_h=9 * mm):
    """Draws a bordered grid table: a shaded bold header row plus ruled data
    rows. Blank cells are left empty, leaving real writable space rather than
    flattening the table into run-on paragraph text."""
    n_cols = len(headers)
    col_w = max_width / n_cols
    header_h = 8 * mm
    y = top_y

    c.setFillColor(HEADER_FILL)
    c.rect(x, y - header_h, max_width, header_h, fill=1, stroke=0)
    c.setFont("Helvetica-Bold", 8.5)
    c.setFillColor(NAVY)
    for i, htext in enumerate(headers):
        if not htext:
            continue
        lines = _cell_lines(htext, "Helvetica-Bold", 8.5, col_w - 4)[:2]
        ty = y - header_h / 2 - (len(lines) - 1) * 4.5 + 3
        for line in lines:
            c.drawString(x + i * col_w + 2, ty, line)
            ty -= 9
    y -= header_h

    c.setFont("Helvetica", 8.5)
    for row in rows:
        c.setFillColor(INK)
        for i, val in enumerate(row):
            if val:
                lines = _cell_lines(val, "Helvetica", 8.5, col_w - 4)[:2]
                ty = y - row_h / 2 - (len(lines) - 1) * 4.5 + 3
                for line in lines:
                    c.drawString(x + i * col_w + 2, ty, line)
                    ty -= 9
        y -= row_h

    total_h = header_h + row_h * len(rows)
    bottom_y = top_y - total_h
    c.setStrokeColor(BORDER)
    c.setLineWidth(0.75)
    c.rect(x, bottom_y, max_width, total_h, fill=0, stroke=1)
    c.line(x, top_y - header_h, x + max_width, top_y - header_h)
    ry = top_y - header_h
    for _ in rows:
        ry -= row_h
        c.line(x, ry, x + max_width, ry)
    for i in range(1, n_cols):
        cx = x + i * col_w
        c.line(cx, top_y, cx, bottom_y)

    return bottom_y - 5 * mm


def _draw_form_fields(c, x, top_y, max_width, fields, line_h=13 * mm):
    """Draws each field label with a ruled blank line beneath it to write on,
    for plan/agreement templates that are a list of labels, not a table."""
    y = top_y
    for label in fields:
        c.setFont("Helvetica-Bold", 9)
        c.setFillColor(TEAL_DARK)
        c.drawString(x, y, label)
        y -= 5 * mm
        c.setStrokeColor(BORDER)
        c.setLineWidth(0.75)
        c.line(x, y, x + max_width, y)
        y -= (line_h - 5 * mm)
    return y


def _draw_checklist(c, x, top_y, max_width, items, row_h=7 * mm):
    """Draws a tick-box list: an empty square before each item."""
    y = top_y
    box = 3.6 * mm
    for item_text in items:
        c.setStrokeColor(BORDER)
        c.setLineWidth(1)
        c.rect(x, y - box, box, box, fill=0, stroke=1)
        c.setFont("Helvetica", 9.5)
        c.setFillColor(INK)
        c.drawString(x + box + 3 * mm, y - box + 0.6 * mm, item_text)
        y -= row_h
    return y - 2 * mm


def _draw_body_map(c, x, top_y, max_width, items, figure_w=42 * mm):
    """The body map as a figure with the tick list beside it.

    A list of words asks a pupil to do the mapping in their head. The printed
    sheet and the screen draw from the same geometry in body_map.py, so a pupil
    filling in a photocopy and one filling it in on the mentor's laptop are
    marking the same picture.
    """
    points = body_map.points_for(items)
    scale = figure_w / body_map.VIEW_W
    fig_h = body_map.VIEW_H * scale
    base_y = top_y - fig_h

    def px(vx):
        return x + vx * scale

    def py(vy):                       # the view grows downwards, the page up
        return base_y + (body_map.VIEW_H - vy) * scale

    c.setStrokeColor(BORDER)
    c.setFillColor(HexColor("#F3F2EC"))
    c.setLineWidth(1)
    for part in body_map.PARTS.values():
        if part[0] == "circle":
            _, vx, vy, r = part
            c.circle(px(vx), py(vy), r * scale, fill=1, stroke=1)
        else:
            _, vx, vy, w, h, r = part
            c.roundRect(px(vx), py(vy + h), w * scale, h * scale, r * scale, fill=1, stroke=1)

    for pt in points:
        if pt["whole"]:
            c.setStrokeColor(TEAL_DARK)
            c.setDash(3, 3)
            c.roundRect(px(8), py(176), 84 * scale, 172 * scale, 14 * scale, fill=0, stroke=1)
            c.setDash()
        c.setFillColor(TEAL_DARK)
        c.setStrokeColor(TEAL_DARK)
        c.circle(px(pt["x"]), py(pt["y"]), body_map.DOT_R * scale, fill=1, stroke=0)
        c.setFillColor(HexColor("#FFFFFF"))
        c.setFont("Helvetica-Bold", 7.5)
        c.drawCentredString(px(pt["x"]), py(pt["y"]) - 2.6, str(pt["n"]))

    # The list sits to the right of the figure, numbered to match.
    list_x = x + figure_w + 8 * mm
    y = top_y - 2 * mm
    box = 3.6 * mm
    for n, item_text in enumerate(items, start=1):
        c.setStrokeColor(BORDER)
        c.setFillColor(HexColor("#FFFFFF"))
        c.setLineWidth(1)
        c.rect(list_x, y - box, box, box, fill=1, stroke=1)
        c.setFillColor(TEAL_DARK)
        c.circle(list_x + box + 4 * mm, y - box + 1.3 * mm, 2.4 * mm, fill=1, stroke=0)
        c.setFillColor(HexColor("#FFFFFF"))
        c.setFont("Helvetica-Bold", 7)
        c.drawCentredString(list_x + box + 4 * mm, y - box + 0.5 * mm, str(n))
        c.setFont("Helvetica", 9.5)
        c.setFillColor(INK)
        c.drawString(list_x + box + 8 * mm, y - box + 0.6 * mm, item_text)
        y -= 8 * mm
    return min(base_y, y) - 4 * mm


def _body_map_height(items, figure_w=42 * mm):
    fig = body_map.VIEW_H * (figure_w / body_map.VIEW_W)
    return max(fig, 8 * mm * len(items)) + 6 * mm


def _table_height(headers, rows, row_h=9 * mm):
    return 8 * mm + row_h * len(rows)


def _form_height(fields, line_h=13 * mm):
    return line_h * len(fields)


def _checklist_height(items, row_h=7 * mm):
    return row_h * len(items)


CYCLE_KEYWORDS = ("cycle diagram", "cycle chart")
SCALE_KEYWORDS = ("thermometer", "scale")


# --- card artwork -----------------------------------------------------------
# The screen and the printed pack draw from one sprite file, templates/_card_art.svg,
# so a motif is never redrawn twice or allowed to drift between the two.
# The motifs are drawn by svgdraw.py, which renders them with reportlab
# directly. svglib was removed: it drew each symbol at its native size instead
# of inside the box it was given, which is what put artwork over the card text.

_ART_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), "templates", "_card_art.svg")
_art_cache = {}


def _art_symbol(art_id):
    """Pulls one <symbol> out of the sprite and returns it as a standalone SVG."""
    if art_id in _art_cache:
        return _art_cache[art_id]
    try:
        with open(_ART_PATH, "r", encoding="utf-8") as f:
            sprite = f.read()
    except OSError:
        _art_cache[art_id] = None
        return None
    m = re.search(r'<symbol id="%s"[^>]*>(.*?)</symbol>' % re.escape(art_id), sprite, re.S)
    if not m:
        _art_cache[art_id] = None
        return None
    _art_cache[art_id] = m.group(1)
    return m.group(1)


# On again. The motifs are now drawn directly with reportlab rather than through
# svglib, which rendered each symbol at its native size instead of inside the box
# it was given — that is what put artwork on top of the card text. The new
# renderer scales explicitly into the box it is handed, and all 100 symbols were
# checked by eye before this was switched back on.
ART_IN_PDF = True


def _draw_art(c, art_id, x, y, width, height):
    """Draws a motif, or quietly does nothing if it can't.

    A missing illustration is a lesser fault than a pack that won't generate, so
    any failure here is swallowed and the card prints as text."""
    if not ART_IN_PDF or not art_id:
        return False
    body = _art_symbol(art_id)
    if not body:
        return False
    try:
        import svgdraw
        return svgdraw.draw_symbol(c, body, x, y, width, height, TEAL_DARK)
    except Exception:
        return False


def _card_height(card):
    """A printed card is a fixed box: title, optional body, artwork, note.

    42mm rather than 30mm: the box was shrunk while the artwork was switched
    off, and the illustration needs its space back."""
    return 42 * mm


def _draw_cut_cards(c, x, y, max_width, cards, new_page=None, bottom=None):
    """Draws a card set as a grid of dashed boxes, two per row, so the sheet can
    be cut up. Each card carries its illustration, drawn from the same sprite the
    screen uses."""
    gap = 4 * mm
    col_w = (max_width - gap) / 2
    card_h = 42 * mm
    for i, card in enumerate(cards):
        col = i % 2
        if col == 0 and i > 0:
            y -= card_h + gap
            # A twelve-card set is taller than a page, so the whole grid can
            # never fit in one block however the estimate is calculated. Break
            # between rows instead, and carry on at the top of the next page.
            if new_page is not None and bottom is not None and y - card_h < bottom:
                y = new_page()
        cx = x + col * (col_w + gap)
        c.setStrokeColor(BORDER)
        c.setLineWidth(0.9)
        c.setDash(3, 2)
        c.rect(cx, y - card_h, col_w, card_h, fill=0, stroke=1)
        c.setDash()

        ty = y - 6 * mm
        cat = _clean_pdf_text(card.get("cat", ""))
        if cat:
            # The category was drawn 1.5mm ABOVE the cursor and the cursor then
            # moved down only 1mm, so the title landed on top of it: a card
            # numbered 1 printed its "1" through "Never tried it".
            c.setFillColor(MUTED)
            c.setFont("Helvetica-Bold", 6.5)
            c.drawString(cx + 3 * mm, ty, cat.upper())
            ty -= 4.5 * mm
        c.setFillColor(TEAL_DARK)
        c.setFont("Helvetica-Bold", 9)
        title = _clean_pdf_text(card.get("title") or card.get("text", ""))
        ty = _wrap(c, title, cx + 3 * mm, ty, col_w - 6 * mm,
                   font="Helvetica-Bold", size=9, leading=11, color=TEAL_DARK)
        body = _clean_pdf_text(card.get("text", "")) if card.get("title") else ""
        if body:
            ty -= 1 * mm
            ty = _wrap(c, body, cx + 3 * mm, ty, col_w - 6 * mm, size=8, leading=10, color=INK)
        _draw_art(c, card.get("art"), cx + 3 * mm, y - card_h + 8 * mm,
                  col_w - 6 * mm, 14 * mm)
        note = _clean_pdf_text(card.get("note", ""))
        if note:
            c.setFillColor(MUTED)
            c.setFont("Helvetica-Oblique", 7)
            c.drawString(cx + 3 * mm, y - card_h + 3 * mm, note)
    rows = (len(cards) + 1) // 2
    return y - card_h - (rows - 1) * 0 if rows == 1 else y - card_h


def _cards_height(cards):
    # Must match _draw_cut_cards: 42mm per card and a 4mm gap between rows. It
    # said 32mm, so a set of six was estimated 30mm shorter than it draws — the
    # page-break check passed and the last row ran off the bottom of the sheet.
    rows = (len(cards) + 1) // 2
    return rows * 42 * mm + (rows - 1) * 4 * mm


# A step's illustration sits to the left of its text, so a sequence reads as a
# row of pictures down the page. Text is inset past the art either way, so a
# step without artwork still lines up with the ones that have it.
_STEP_ART_W = 16 * mm
_STEP_ART_H = 12 * mm


def _draw_steps(c, x, y, max_width, steps):
    """Numbered steps with their illustrations, kept together as a sequence.

    The artwork was missing from printed packs entirely: the screen drew each
    step's motif and this did not, so a diagram like the nicotine cycle printed
    as a list of sentences with the picture that explains it left out."""
    text_x = x + _STEP_ART_W + 3 * mm
    for i, st in enumerate(steps, start=1):
        top = y
        c.setFillColor(TEAL_DARK)
        c.setFont("Helvetica-Bold", 9)
        title = _clean_pdf_text(st.get("title", ""))
        c.drawString(text_x, y, f"{i}. {title}")
        y -= 4.5 * mm
        text = _clean_pdf_text(st.get("text", ""))
        if text:
            y = _wrap(c, text, text_x, y, max_width - (text_x - x), size=8.5,
                      leading=10.5, color=INK)

        # Centre the motif against the block of text it belongs to.
        used = top - y
        art_y = top - max(used, _STEP_ART_H) + (max(used, _STEP_ART_H) - _STEP_ART_H) / 2
        _draw_art(c, st.get("art"), x, art_y, _STEP_ART_W, _STEP_ART_H)
        if used < _STEP_ART_H:
            y = top - _STEP_ART_H
        y -= 3.5 * mm
    return y


def _steps_height(steps):
    # Each step is at least as tall as its artwork.
    return len(steps) * 18 * mm


def resource_pack_pdf(course_num, course_title, items):
    """
    Generates the full downloadable resource pack for a course, containing
    every handout/card/template referenced by that course's sessions.
    course_num: two-digit string, e.g. "01"
    course_title: str
    items: list of {"name": str, "body": str}
    """
    path = os.path.join(PDF_DIR, f"resource_pack_{course_num}.pdf")
    w, h = A4
    c = canvas.Canvas(path, pagesize=A4)
    margin = 18 * mm
    max_width = w - margin * 2
    state = {"y": h - margin}

    def header():
        state["y"] = h - margin
        c.setFillColor(NAVY)
        c.setFont("Helvetica-Bold", 16)
        c.drawString(margin, state["y"], "Phil - Resource Pack")
        state["y"] -= 7 * mm
        c.setFillColor(MUTED)
        c.setFont("Helvetica", 10)
        c.drawString(margin, state["y"], _clean_pdf_text(course_title))
        state["y"] -= 6 * mm
        c.setStrokeColor(BORDER)
        c.setLineWidth(0.75)
        c.line(margin, state["y"], w - margin, state["y"])
        state["y"] -= 9 * mm

    def new_page():
        c.showPage()
        header()

    def cut_line():
        """A dashed rule between assets, so a pack can be cut into separate
        sheets. Drawn between items only, never at the top or bottom of a page:
        a cut mark with nothing under it invites someone to cut off a margin."""
        state["y"] -= 7 * mm
        c.setStrokeColor(BORDER)
        c.setLineWidth(0.6)
        c.setDash(2, 3)
        c.line(margin, state["y"], w - margin, state["y"])
        c.setDash()
        c.setFillColor(MUTED)
        c.setFont("Helvetica", 6.5)
        c.drawRightString(w - margin, state["y"] + 1.6 * mm, "cut here")
        state["y"] -= 9 * mm

    header()

    SEPARATOR_H = 16 * mm

    for index, item in enumerate(items):
        first_on_page = state["y"] >= h - margin - 26 * mm
        name = item.get("name", "")
        name_lower = name.lower()
        table = item.get("table")
        form = item.get("form")
        checklist = item.get("checklist")
        is_cycle = any(k in name_lower for k in CYCLE_KEYWORDS)
        is_scale = any(k in name_lower for k in SCALE_KEYWORDS)

        cards = item.get("cards")
        steps = item.get("steps")
        if cards:
            # An item can carry a table and cards together, the same way one can
            # carry a checklist and a form. Reserve room for both or the table
            # gets pushed off the page it was measured for.
            needed = 24 * mm + _cards_height(cards)
            if table:
                needed += _table_height(table["headers"], table["rows"])
        elif steps:
            needed = 24 * mm + _steps_height(steps)
            if table:
                needed += _table_height(table["headers"], table["rows"])
        elif table:
            needed = 24 * mm + _table_height(table["headers"], table["rows"])
        elif checklist and item.get("figure") == "body-map":
            needed = 24 * mm + _body_map_height(checklist["items"])
        elif checklist:
            # An item can carry a checklist and a form together. The chain drew
            # only one, so eleven of "How staff can help me"'s thirteen lines
            # never printed.
            needed = 24 * mm + _checklist_height(checklist["items"])
            if form:
                needed += _form_height(form["fields"])
        elif form:
            needed = 24 * mm + _form_height(form["fields"])
        elif is_cycle:
            needed = 140 * mm
        elif is_scale:
            needed = 55 * mm
        else:
            needed = 28 * mm
        if state["y"] < margin + needed + SEPARATOR_H:
            new_page()
            first_on_page = True

        # Separate this asset from the previous one, unless it starts a page.
        if index > 0 and not first_on_page:
            cut_line()

        c.setFillColor(TEAL_DARK)
        c.setFont("Helvetica-Bold", 12)
        c.drawString(margin, state["y"], _clean_pdf_text(name))
        state["y"] -= 7 * mm

        body = _clean_pdf_text(item.get("body", ""))

        if cards:
            for para in body.split("\n"):
                state["y"] = _wrap(c, para, margin, state["y"], max_width, font="Helvetica", size=9.5, leading=13, color=INK)
            state["y"] -= 4 * mm
            if table:
                # The chain used to be `if cards ... elif table`, so any item
                # with both printed the cards and silently dropped the table.
                # On screen the table comes first, so it does here too.
                state["y"] = _draw_grid_table(c, margin, state["y"], max_width,
                                              table["headers"], table["rows"])
                state["y"] -= 4 * mm
            state["y"] = _draw_cut_cards(c, margin, state["y"], max_width, cards,
                                         new_page=lambda: (new_page(), state["y"])[1],
                                         bottom=margin)
            state["y"] -= 3 * mm

        elif steps:
            for para in body.split("\n"):
                state["y"] = _wrap(c, para, margin, state["y"], max_width, font="Helvetica", size=9.5, leading=13, color=INK)
            state["y"] -= 4 * mm
            if table:
                state["y"] = _draw_grid_table(c, margin, state["y"], max_width,
                                              table["headers"], table["rows"])
                state["y"] -= 4 * mm
            state["y"] = _draw_steps(c, margin, state["y"], max_width, steps)
            if item.get("note"):
                c.setFillColor(MUTED)
                c.setFont("Helvetica-Oblique", 7.5)
                c.drawString(margin, state["y"], _clean_pdf_text(item["note"]))
                state["y"] -= 5 * mm

        elif table:
            for para in body.split("\n"):
                state["y"] = _wrap(c, para, margin, state["y"], max_width, font="Helvetica", size=9.5, leading=13, color=INK)
            state["y"] -= 4 * mm
            state["y"] = _draw_grid_table(c, margin, state["y"], max_width, table["headers"], table["rows"])

        elif checklist:
            for para in body.split("\n"):
                state["y"] = _wrap(c, para, margin, state["y"], max_width, font="Helvetica", size=9.5, leading=13, color=INK)
            state["y"] -= 4 * mm
            if item.get("figure") == "body-map":
                state["y"] = _draw_body_map(c, margin, state["y"], max_width, checklist["items"])
            else:
                state["y"] = _draw_checklist(c, margin, state["y"], max_width, checklist["items"])
            if form:
                state["y"] -= 3 * mm
                state["y"] = _draw_form_fields(c, margin, state["y"], max_width, form["fields"])

        elif form:
            for para in body.split("\n"):
                state["y"] = _wrap(c, para, margin, state["y"], max_width, font="Helvetica", size=9.5, leading=13, color=INK)
            state["y"] -= 4 * mm
            state["y"] = _draw_form_fields(c, margin, state["y"], max_width, form["fields"])
            state["y"] -= 3 * mm

        elif is_cycle:
            steps = _parse_numbered_steps(body)
            intro = body.split("\n")[0] if body else ""
            if intro and not _STEP_RE.match(intro.strip()):
                state["y"] = _wrap(c, intro, margin, state["y"], max_width, font="Helvetica", size=9.5, leading=13, color=INK)
                state["y"] -= 3 * mm
            if steps:
                diagram_h = 76 * mm
                cx = margin + max_width / 2
                cy = state["y"] - 37 * mm
                radius = 22 * mm
                _draw_cycle_diagram(c, cx, cy, radius, steps)
                state["y"] -= diagram_h + 6 * mm
                for step in steps:
                    state["y"] = _wrap(c, step["full"], margin, state["y"], max_width, font="Helvetica", size=8.5, leading=11.5, color=MUTED)
                    state["y"] -= 3 * mm
            else:
                for para in body.split("\n"):
                    state["y"] = _wrap(c, para, margin, state["y"], max_width, font="Helvetica", size=9.5, leading=13, color=INK)
                    state["y"] -= 7 * mm

        elif is_scale:
            points = _parse_scale_points(body)
            intro_lines = [ln for ln in body.split("\n") if not _SCALE_RE.match(ln.strip())]
            for ln in intro_lines:
                state["y"] = _wrap(c, ln, margin, state["y"], max_width, font="Helvetica", size=9.5, leading=13, color=INK)
            state["y"] -= 4 * mm
            if points:
                _draw_scale_diagram(c, margin, state["y"], max_width, points)
                state["y"] -= 22 * mm
            state["y"] -= 4 * mm

        else:
            for para in body.split("\n"):
                if state["y"] < margin + 14 * mm:
                    new_page()
                state["y"] = _wrap(c, para, margin, state["y"], max_width, font="Helvetica", size=9.5, leading=13, color=INK)
            state["y"] -= 7 * mm

        # Where a resource teaches an established technique, say where it comes
        # from. Printed once under the resource, above the cut line, so it stays
        # on the mentor's sheet rather than on every card a pupil handles.
        if item.get("source"):
            if state["y"] < margin + 12 * mm:
                new_page()
            state["y"] = _wrap(c, "Source: " + _clean_pdf_text(item["source"]),
                               margin, state["y"], max_width,
                               font="Helvetica-Oblique", size=7.5, leading=9.5,
                               color=MUTED)
            state["y"] -= 4 * mm

    c.showPage()
    c.save()
    return path


def legal_doc_pdf(key, paras):
    path = os.path.join(PDF_DIR, f"legal_{key}.pdf")
    w, h = A4
    c = canvas.Canvas(path, pagesize=A4)
    margin = 20 * mm
    max_width = w - margin * 2
    state = {"y": h - margin}

    def new_page():
        c.showPage()
        state["y"] = h - margin

    cover = []
    body = paras
    for i, item in enumerate(paras):
        if item[0] == "h1":
            cover = paras[:i]
            body = paras[i:]
            break

    if cover:
        kicker = cover[0][1] if len(cover) > 0 else ""
        title = cover[1][1] if len(cover) > 1 else ""
        meta_lines = [t for s, t in cover[2:] if s != "blank" and t]

        c.setFillColor(TEAL_DARK)
        c.setFont("Helvetica-Bold", 9.5)
        c.drawString(margin, state["y"], _clean_pdf_text(kicker.upper()))
        state["y"] -= 9 * mm

        c.setFillColor(NAVY)
        c.setFont("Helvetica-Bold", 20)
        title_lines = simpleSplit(_clean_pdf_text(title), "Helvetica-Bold", 20, max_width)
        for line in title_lines:
            c.drawString(margin, state["y"], line)
            state["y"] -= 8.5 * mm
        state["y"] -= 2 * mm

        for line in meta_lines:
            state["y"] = _wrap(c, line, margin, state["y"], max_width, font="Helvetica", size=10, leading=14, color=MUTED)
        state["y"] -= 4 * mm

        c.setStrokeColor(BORDER)
        c.setLineWidth(0.75)
        c.line(margin, state["y"], w - margin, state["y"])
        state["y"] -= 10 * mm

    for style, text in body:
        text = _clean_pdf_text(text)
        if style == "blank":
            state["y"] -= 3 * mm
            continue
        if style == "h1":
            if state["y"] < margin + 24 * mm:
                new_page()
            state["y"] -= 4 * mm
            c.setFillColor(TEAL_DARK)
            c.setFont("Helvetica-Bold", 13.5)
            lines = simpleSplit(text, "Helvetica-Bold", 13.5, max_width)
            for line in lines:
                if state["y"] < margin + 14 * mm:
                    new_page()
                c.drawString(margin, state["y"], line)
                state["y"] -= 6.2 * mm
            state["y"] -= 2 * mm
        elif style == "h2":
            if state["y"] < margin + 18 * mm:
                new_page()
            state["y"] -= 2 * mm
            c.setFillColor(NAVY)
            c.setFont("Helvetica-Bold", 11)
            lines = simpleSplit(text, "Helvetica-Bold", 11, max_width)
            for line in lines:
                if state["y"] < margin + 12 * mm:
                    new_page()
                c.drawString(margin, state["y"], line)
                state["y"] -= 5.4 * mm
            state["y"] -= 1.5 * mm
        elif style == "li":
            if state["y"] < margin + 12 * mm:
                new_page()
            bullet_indent = 4 * mm
            c.setFont("Helvetica", 9.5)
            c.setFillColor(INK)
            c.drawString(margin, state["y"], "-")
            lines = simpleSplit(text, "Helvetica", 9.5, max_width - bullet_indent)
            for line in lines:
                if state["y"] < margin + 10 * mm:
                    new_page()
                c.drawString(margin + bullet_indent, state["y"], line)
                state["y"] -= 12.5
            state["y"] -= 1 * mm
        else:
            if state["y"] < margin + 12 * mm:
                new_page()
            state["y"] = _wrap(c, text, margin, state["y"], max_width, font="Helvetica", size=9.5, leading=13, color=INK)
            state["y"] -= 2 * mm

    c.showPage()
    c.save()
    return path
